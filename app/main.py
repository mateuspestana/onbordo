from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import re
import unicodedata

from openpyxl import load_workbook
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware


APP_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = APP_ROOT.parent
EXCEL_PATH = PROJECT_ROOT / "base_barcos_dummy.xlsx"


app = FastAPI(title="OnBordo - API de Veleiros", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


_CACHE: Dict[str, Any] = {"df": None, "services_column": None, "alias_map": None}


def _normalize_name(name: Any) -> str:
    s = str(name).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"R\s*\$", "RS", s, flags=re.IGNORECASE)
    s = re.sub(r"[^A-Za-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_").lower()


def _normalize_services_cell(cell: Any) -> List[str]:
    if cell is None:
        return []
    if isinstance(cell, list):
        return [str(x).strip() for x in cell if str(x).strip()]
    parts = [p.strip() for p in str(cell).replace(";", ",").split(",")]
    return [p for p in parts if p]


def _load_dataframe(refresh: bool = False) -> Tuple[List[Dict[str, Any]], Optional[str], Dict[str, str]]:
    global _CACHE
    if _CACHE["df"] is not None and not refresh:
        return list(_CACHE["df"]), _CACHE["services_column"], dict(_CACHE["alias_map"])  # type: ignore

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado em: {EXCEL_PATH}")

    # Carrega Excel com openpyxl
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # Extrai dados como lista de dicionários
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia")
    
    headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(rows[0])]
    data_rows = rows[1:]
    
    # Converte para lista de dicionários com ID
    records = []
    for i, row in enumerate(data_rows):
        record = {"id": i}
        for j, value in enumerate(row):
            if j < len(headers):
                record[headers[j]] = value
        records.append(record)

    # Constrói mapa de apelidos
    alias_map: Dict[str, str] = {}
    for col in headers:
        alias = _normalize_name(col)
        alias_map.setdefault(alias, col)

    preferred_aliases: Dict[str, str] = {
        "id_do_barco": "ID do Barco",
        "preco_por_dia_rs": "Preço por Dia (R$)",
        "nome_do_barco": "Nome do Barco",
        "marina_porto": "Marina/Porto",
        "pes": "Pés",
        "tripulante": "Tripulante",
        "preco_do_arrais_rs": "Preço do Arrais (R$)",
        "outros_servicos": "Outros Serviços",
    }

    for alias, desired_original in preferred_aliases.items():
        if desired_original in headers:
            alias_map[alias] = desired_original
        else:
            matched = next((c for c in headers if _normalize_name(c) == alias), None)
            if matched is not None:
                alias_map[alias] = matched

    # Detecta coluna de serviços
    service_alias_targets = {"servicos", "servico", "servicos_outros", "outros_servicos", "services"}
    services_col: Optional[str] = None
    for alias, original in alias_map.items():
        if alias in service_alias_targets:
            services_col = original
            break

    # Adiciona services_list processada
    for record in records:
        if services_col and services_col in record:
            record["services_list"] = _normalize_services_cell(record[services_col])
        else:
            record["services_list"] = []

    _CACHE = {"df": records, "services_column": services_col, "alias_map": alias_map}
    return list(records), services_col, alias_map


def _to_records(records: List[Dict[str, Any]], include_services_list: bool = False) -> List[Dict[str, Any]]:
    result = [dict(rec) for rec in records]
    if not include_services_list:
        for rec in result:
            rec.pop("services_list", None)
    return result


def _try_parse_number(value: str) -> Any:
    try:
        if "." in value or "," in value:
            return float(value.replace(",", "."))
        return int(value)
    except Exception:
        return value


def _parse_in_list(value: str) -> List[str]:
    return [v.strip() for v in value.split(",") if v.strip()]


def _apply_generic_filters(records: List[Dict[str, Any]], qp: Dict[str, str], services_col: Optional[str], alias_map: Dict[str, str]) -> List[Dict[str, Any]]:
    reserved = {
        "limit", "offset", "sort_by", "sort_order", "columns",
        "services_any", "services_all", "services_not", "refresh", "format",
    }

    filtered = list(records)

    # Aplica filtros de coluna
    for raw_key, raw_value in qp.items():
        if raw_key in reserved:
            continue

        if "__" in raw_key:
            column, op = raw_key.split("__", 1)
        else:
            column, op = raw_key, "auto"

        # Resolve apelido
        if column not in (filtered[0].keys() if filtered else []):
            alias = _normalize_name(column)
            if alias in alias_map:
                column = alias_map[alias]
            else:
                raise HTTPException(status_code=400, detail=f"Coluna desconhecida para filtro: {column}")

        val = raw_value

        # Aplica filtro
        new_filtered = []
        for record in filtered:
            cell_value = record.get(column)
            
            if op == "auto":
                # Tenta igualdade numérica primeiro, senão contains
                try:
                    comp = _try_parse_number(val)
                    if isinstance(cell_value, (int, float)) and cell_value == comp:
                        new_filtered.append(record)
                    elif str(cell_value).lower().find(str(val).lower()) >= 0:
                        new_filtered.append(record)
                except:
                    if str(cell_value).lower().find(str(val).lower()) >= 0:
                        new_filtered.append(record)
            elif op == "eq":
                if isinstance(cell_value, (int, float)):
                    comp = _try_parse_number(val)
                    if cell_value == comp:
                        new_filtered.append(record)
                else:
                    if str(cell_value).lower() == str(val).lower():
                        new_filtered.append(record)
            elif op == "contains":
                if str(cell_value).lower().find(str(val).lower()) >= 0:
                    new_filtered.append(record)
            elif op == "in":
                values = _parse_in_list(val)
                if isinstance(cell_value, (int, float)):
                    parsed = [_try_parse_number(v) for v in values]
                    if cell_value in parsed:
                        new_filtered.append(record)
                else:
                    lowered = [v.lower() for v in values]
                    if str(cell_value).lower() in lowered:
                        new_filtered.append(record)
            elif op in {"lt", "lte", "gt", "gte"}:
                try:
                    comp = _try_parse_number(val)
                    if isinstance(cell_value, (int, float)):
                        if op == "lt" and cell_value < comp:
                            new_filtered.append(record)
                        elif op == "lte" and cell_value <= comp:
                            new_filtered.append(record)
                        elif op == "gt" and cell_value > comp:
                            new_filtered.append(record)
                        elif op == "gte" and cell_value >= comp:
                            new_filtered.append(record)
                except:
                    pass
            elif op == "between":
                parts = _parse_in_list(val)
                if len(parts) != 2:
                    raise HTTPException(status_code=400, detail=f"Filtro between inválido para {column}. Use min,max")
                try:
                    lo, hi = _try_parse_number(parts[0]), _try_parse_number(parts[1])
                    if isinstance(cell_value, (int, float)) and lo <= cell_value <= hi:
                        new_filtered.append(record)
                except:
                    pass
            elif op == "isnull":
                truthy = str(val).lower() in {"1", "true", "t", "yes", "y"}
                is_null = cell_value is None or (isinstance(cell_value, str) and not cell_value.strip())
                if (truthy and is_null) or (not truthy and not is_null):
                    new_filtered.append(record)
            else:
                raise HTTPException(status_code=400, detail=f"Operador desconhecido: {op} (coluna {column})")
        
        filtered = new_filtered

    # Filtros de serviços
    any_q = qp.get("services_any")
    all_q = qp.get("services_all")
    not_q = qp.get("services_not")

    if any_q or all_q or not_q:
        def norm_list(v: Optional[str]) -> List[str]:
            if not v:
                return []
            return [s.strip().lower() for s in _parse_in_list(v)]

        any_list = set(norm_list(any_q))
        all_list = set(norm_list(all_q))
        not_list = set(norm_list(not_q))

        def matches(row_services: Iterable[str]) -> bool:
            sset = {s.strip().lower() for s in row_services if s}
            if any_list and sset.isdisjoint(any_list):
                return False
            if all_list and not all_list.issubset(sset):
                return False
            if not_list and not sset.isdisjoint(not_list):
                return False
            return True

        filtered = [r for r in filtered if matches(r.get("services_list", []))]

    return filtered


@app.get("/")
def root() -> Dict[str, Any]:
    return {
        "name": app.title,
        "version": app.version,
        "docs": "/docs",
        "openapi": "/openapi.json",
        "endpoints": {"boats": "/boats", "schema": "/schema"},
    }


@app.get("/schema")
def schema(refresh: bool = False) -> Dict[str, Any]:
    records, services_col, alias_map = _load_dataframe(refresh=refresh)
    if records:
        cols = [{"name": str(c), "dtype": "mixed"} for c in records[0].keys() if c != "services_list"]
    else:
        cols = []
    aliases = [{"alias": a, "column": o} for a, o in alias_map.items() if o != "services_list"]
    return {"columns": cols, "aliases": aliases, "services_column": services_col, "count": len(records)}


@app.get("/boats")
def list_boats(request: Request) -> Dict[str, Any]:
    qp = dict(request.query_params)

    refresh = str(qp.get("refresh", "false")).lower() in {"1", "true", "t", "yes", "y"}
    records, services_col, alias_map = _load_dataframe(refresh=refresh)

    filtered = _apply_generic_filters(records, qp, services_col, alias_map)

    # Ordenação
    sort_by = qp.get("sort_by")
    if sort_by and filtered:
        sort_keys = []
        for token in sort_by.split(","):
            token = token.strip()
            if not token:
                continue
            if token.startswith("-"):
                col = token[1:]
                reverse = True
            elif token.startswith("+"):
                col = token[1:]
                reverse = False
            else:
                col = token
                reverse = False
            
            # Resolve apelido
            if col not in filtered[0].keys():
                alias = _normalize_name(col)
                if alias in alias_map:
                    col = alias_map[alias]
                else:
                    raise HTTPException(status_code=400, detail=f"Coluna inexistente para ordenação: {col}")
            
            sort_keys.append((col, reverse))
        
        # Aplica ordenação múltipla
        for col, reverse in reversed(sort_keys):
            filtered.sort(key=lambda x: x.get(col, ""), reverse=reverse)

    # Projeção de colunas
    columns = qp.get("columns")
    if columns and filtered:
        requested = [c.strip() for c in columns.split(",") if c.strip()]
        if "id" not in requested:
            requested = ["id"] + requested
        
        resolved: List[str] = []
        missing: List[str] = []
        for c in requested:
            if c in filtered[0].keys():
                resolved.append(c)
            elif c == "services_list":
                continue
            else:
                alias = _normalize_name(c)
                if alias in alias_map:
                    resolved.append(alias_map[alias])
                else:
                    missing.append(c)
        
        if missing:
            raise HTTPException(status_code=400, detail=f"Colunas inexistentes na projeção: {', '.join(missing)}")
        
        # Projeta colunas
        filtered = [{k: record[k] for k in resolved if k in record} for record in filtered]

    # Paginação
    try:
        limit = int(qp.get("limit", "0"))
        offset = int(qp.get("offset", "0"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Parâmetros de paginação inválidos")

    total = len(filtered)
    if offset > 0:
        filtered = filtered[offset:]
    if limit and limit > 0:
        filtered = filtered[:limit]

    include_services_list = str(qp.get("format", "")).lower() == "debug"
    data = _to_records(filtered, include_services_list=include_services_list)
    return {"total": total, "count": len(data), "items": data}


@app.get("/boats/{boat_id}")
def get_boat(boat_id: int, refresh: bool = False) -> Dict[str, Any]:
    records, _, _ = _load_dataframe(refresh=refresh)
    for record in records:
        if record.get("id") == boat_id:
            return _to_records([record])[0]
    raise HTTPException(status_code=404, detail="Barco não encontrado")


